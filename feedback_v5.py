
#!/usr/bin/env python3
"""
stability-feedback v5.3.3 放样反馈处理
========================================
功能：
  1. 接收放样反馈（药品/批号/实际放样日期）
  2. 读取对应药品的稳定性研究计划.xlsx
  3. 解析各条件区块的时间点：
     - 天数字段：原始值即为距放样天数，标签显示原始值（如"10"），
       计划日期 = 放样日期 + 该原始值（天数）
     - 月历列（1月/2月等）：标签保留月历标签（如"1月"），
       计划日期 = 放样日期 + relativedelta(months=月数)
     - 两类点按实际距放样天数统一排序，生成时间轴顺序
  4. 以实际放样日期为基准，计算各节点计划取样日期
  5. 检测项用量：✓=0（检测但不消耗样品），-=0（不检测），数字=该数字
  6. 写入「取样计划总表.xlsx」，列顺序按要求调整，删除备注列

用法：
  python feedback_v5.py --drug "黄体酮注射液" \\
    --placement-date "2026-04-09" [--batch "LM-xxx"] [--condition "长期25℃"] \\
    [--notes "无异常"] [--dry-run]
"""

import argparse
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from dateutil.relativedelta import relativedelta
except ImportError:
    print("❌ 需要 python-dateutil: pip install python-dateutil")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule, FormulaRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 需要 openpyxl: pip install openpyxl")
    sys.exit(1)


# ============ 路径配置 ============
AGENT_ROOT = Path("/home/wangyc/.openclaw/workspace/agents/auto-formula-scheme")
PLAN_ROOT  = AGENT_ROOT / "stability_plans"
MASTER_TABLE = PLAN_ROOT / "取样计划总表.xlsx"
# =================================

# 批号前缀别名映射：计划文件批号前缀 -> 实际批号前缀
# 用于处理计划文件批号与实际批号前缀不一致的情况
BATCH_PREFIX_ALIAS = {
    "LM002-":   "LMS002-",  # 黄体酮预灌封：计划LM002- vs 实际LMS002-
    "LM2002-":  "LMS002-",  # 黄体酮预灌封：计划LM2002- vs 实际LMS002-
}

def normalize_batch_key(batch):
    """
    提取批号的标准化比较键：去除常见前缀变体，保留"年-月日"比较键。
    例如：LM002-260417-01, LM2002-260417-01, LMS002-260417-01
         均被规范化为               2604-17-01（年2位+月 + - + 日 + - + 序号）
    日期格式：YYYYMMDD（如 260417 = 2026-04-17）
    """
    import re
    # 匹配：可选 LM/LMS + 可选数字串- + 年月日(6位) + - + 序号
    # 例：LMS002-260417-01 → 规范化键：2604-17-01
    m = re.match(r'^LM(S)?\d*-(\d{6})-(\d+)$', batch)
    if m:
        _s = m.group(1)          # optional S (unused)
        date_part = m.group(2)   # 260417
        seq = m.group(3)         # 01
        year2 = date_part[:2]    # 26
        month  = date_part[2:4]  # 04
        day    = date_part[4:6]  # 17
        return f"{year2}{month}-{day}-{seq}"
    return batch  # 无法解析时返回原值

def normalize_batch_to_plan(batch):
    """将实际批号转换为计划文件格式（如果需要）。"""
    for plan_prefix, actual_prefix in BATCH_PREFIX_ALIAS.items():
        if batch.startswith(actual_prefix):
            return batch.replace(actual_prefix, plan_prefix, 1)
    return batch

def normalize_batch_to_actual(batch):
    """将计划文件批号转换为实际格式。"""
    for plan_prefix, actual_prefix in BATCH_PREFIX_ALIAS.items():
        if batch.startswith(plan_prefix):
            return batch.replace(plan_prefix, actual_prefix, 1)
    return batch


# 检测项目（按常见顺序，可扩展）
DETECTION_ITEMS = ["外观/性状", "有关物质", "pH值", "不溶性微粒", "含量", "无菌", "细菌内毒素"]

DATE_FMT = "YYYY-MM-DD"

# ---- Excel 样式常量（模块级别，供所有函数共享）----
_SIDE_THIN   = Side(style="thin",   color="CCCCCC")
CELL_BORDER  = Border(left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=_SIDE_THIN)
CELL_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
HDR_FILL     = PatternFill(fill_type="solid", fgColor="1F4E79")
HDR_FONT     = Font(bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(size=10)
ZEBRA_ODD    = PatternFill(fill_type="solid", fgColor="FFFFFF")
ZEBRA_EVEN   = PatternFill(fill_type="solid", fgColor="EEF3F8")
STATUS_EXEC  = (
    PatternFill(fill_type="solid", fgColor="DDEEFF"),
    Font(color="1F4E79", bold=True),
)
STATUS_DONE  = (
    PatternFill(fill_type="solid", fgColor="D5F5E3"),
    Font(color="1E8449", bold=True),
)
STATUS_OVER  = (
    PatternFill(fill_type="solid", fgColor="FADBD8"),
    Font(color="C0392B", bold=True),
)


def normalize_date(s: str) -> str:
    """将各种日期格式统一为 YYYY-MM-DD"""
    if not s:
        return ""
    s = s.strip().replace("年", "-").replace("月", "-").replace("日", "").replace("/", "-")
    parts = re.split(r"[-]", s)
    if len(parts) >= 3:
        return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    return s


def to_num(v):
    """将数值或字符串数值转为 float，否则返回 None"""
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except (ValueError, AttributeError):
            return None
    return None


def find_plan_file(drug_name: str) -> Path:
    """在 stability_plans/{药品名}/ 下查找 *_稳定性计划.xlsx，优先选择包含最多批号的文件"""
    candidates = []
    for item in PLAN_ROOT.iterdir():
        if item.is_dir() and drug_name in item.name:
            for f in item.glob("*_稳定性计划.xlsx"):
                candidates.append(f)
    if not candidates:
        return None
    # 按文件名长度降序排列（更具体的文件名优先），避免 --xxx 占位文件被误选
    candidates.sort(key=lambda f: len(f.name), reverse=True)
    return candidates[0]


def resolve_plan_path(drug_name: str, plan_file_arg: str = None) -> Path:
    """解析计划文件路径：优先使用 --plan-file 参数，否则用 find_plan_file 自动查找"""
    if plan_file_arg:
        p = Path(plan_file_arg)
        if p.is_absolute():
            return p
        # 相对路径：相对于 PLAN_ROOT
        return PLAN_ROOT / p
    return find_plan_file(drug_name)


def parse_plan(plan_path: Path, target_batches: list = None):
    """
    解析稳定性研究计划.xlsx，返回:
    {
        "drug_name": str,
        "batches": [str],
        "records": {
            "条件名": [(时间点label, 距放样天数, {检测项:用量}, 总计), ...]
        }
    }

    关键设计：
    - 月历列（1月/2月等）和纯天数列分开处理，互不去重
    - 月历列使用计划表里的实际天数（如3月=91天），不用 month_map 估算
    - 纯天数列（10天/20天/40天等）全部保留
    - ✓ → 消耗量=0（检测但不消耗样品）
    - '-' → 消耗量=0（不检测）
    - 数字 → 消耗量=该数字
    """
    wb = openpyxl.load_workbook(plan_path, data_only=True)
    ws = wb["稳定性研究计划"]

    # ---- 基本信息：药品名 ----
    drug_name = ""
    for r in range(1, 10):
        for c in range(1, 10):
            v = ws.cell(r, c).value
            if v and "药品" in str(v) and c < 9:
                drug_name = str(ws.cell(r, c + 1).value or "").strip()

    # ---- 批号 ----
    batches = []
    batch_row = None
    for r in range(1, 10):
        for c in range(1, 10):
            v = ws.cell(r, c).value
            if v and "批号" in str(v):
                batch_row = r
                break
    if batch_row:
        for r in range(batch_row, min(batch_row + 20, 30)):
            for c in range(1, 10):
                bv = ws.cell(r, c).value
                # 拒绝纯日期(YYYY-MM-DD)和纯文字；只接受含LM或以数字开头后跟日期的
                if bv:
                    bv_str = str(bv).strip()
                    # 排除：纯数字、纯文字（无字母无横杠）、日期格式 YYYY-MM-DD
                    if bv_str.isdigit():
                        continue
                    if "-" not in bv_str and "LM" not in bv_str:
                        continue
                    # 排除纯日期如 2026-04-22（以4个数字开头，中间有-）
                    import re
                    if re.match(r"^\d{4}-\d{2}-\d{2}$", bv_str):
                        continue
                    batches.append(bv_str)

    batch_alias_forward = {}  # plan_batch -> actual_batch
    if target_batches:
        # 建立目标批号的标准化键索引
        target_keys = {normalize_batch_key(b): b for b in target_batches}
        # 用标准化键匹配，保留匹配的批号及其原始值
        plan_key_to_batch = {normalize_batch_key(b): b for b in batches}
        common_keys = set(target_keys.keys()) & set(plan_key_to_batch.keys())
        matched_plan_batches = [plan_key_to_batch[k] for k in common_keys]
        matched_actual_batches = [target_keys[k] for k in common_keys]
        batches = matched_plan_batches
        # 建立反向映射
        for plan_b, actual_b in zip(matched_plan_batches, matched_actual_batches):
            if plan_b != actual_b:
                batch_alias_forward[plan_b] = actual_b

    # 将批号转换为实际格式（供写入总表用）
    actual_batches = [batch_alias_forward.get(b, b) for b in batches]

    # ---- 条件区块定义：(条件名, header_row, day_row, month_row, data_start, end_row) ----
    # 注：以下行号为该计划文件的实际位置（阴凉行21，检测项目行22；高温行29，检测项目行30）
    blocks = [
        ("长期25℃",  12, 13, 14, 15, 18),
        ("阴凉20℃",  21, 22, 23, 24, 27),
        ("加速40℃",  28, 29, 30, 31, 34),
        ("高温60℃",  29, 30, 31, 32, 35),
    ]

    result = {}

    for cond_name, hdr_row, day_row_num, month_row_num, data_start, end_row in blocks:
        day_row   = [ws.cell(day_row_num,   c).value for c in range(1, 25)]
        month_row = [ws.cell(month_row_num, c).value for c in range(1, 25)]

        cal_cols = {}   # col_idx(1-based) → (label, months_count)  月历列
        day_cols = {}   # col_idx(1-based) → actual_days               纯天数列

        for col_idx in range(1, 25):
            d_val   = day_row[col_idx - 1]
            m_val   = month_row[col_idx - 1]
            is_month = isinstance(m_val, str) and m_val.endswith("月")
            num     = to_num(d_val)

            if num is None or num < 0:
                continue

            if is_month:
                # 从月份标签提取月数："3月" → 3，"12月" → 12
                month_count = int(re.sub(r"\D", "", m_val))
                cal_cols[col_idx] = (m_val, month_count)
            # 注意：同一列可能同时有 day 和 month（如 col6 day=30 + 1月）
            # 两者都要录入，分别生成独立记录
            if num > 0:
                day_cols[col_idx] = int(num)

        # 构建 time_points（统一入口，按 column index 顺序逐个处理月历列和天数列）
        # placement_ref 仅用于月历点算 actual_days
        placement_ref = datetime(2026, 4, 10)

        all_cols = sorted(set(list(cal_cols.keys()) + list(day_cols.keys())))
        unified = []
        prev_raw = 0       # 上一个 entry 的 day_cols 原始值
        prev_actual = 0   # 上一个 entry 的 actual_days（month 或 day）

        for col_idx in all_cols:
            if col_idx in cal_cols:
                # 月历点
                label, months = cal_cols[col_idx]
                actual_days = (placement_ref + relativedelta(months=months) - placement_ref).days
                unified.append((col_idx, 'month', label, actual_days, months))
                prev_actual = actual_days
                # 如果月历列也包含天数，则更新 prev_raw
                if col_idx in day_cols and day_cols[col_idx] > 0:
                    prev_raw = day_cols[col_idx] # 覆盖前一个 day 的 raw，从月历点起算
            elif col_idx in day_cols and day_cols[col_idx] > 0:
                raw = day_cols[col_idx]
                iv = raw - prev_raw if prev_raw > 0 else raw
                ad = prev_actual + iv
                unified.append((col_idx, 'day', str(raw), ad, iv)) # Label now shows raw value
                prev_raw = raw
                prev_actual = ad

        # 按 actual_days 统一排序
        unified.sort(key=lambda x: (x[3], 0 if x[1]=='day' else 1))

        # col_idx → (label, interval, is_month, actual_days)
        col_info = {}  # col_idx → (label, interval, is_month, actual_days)
        for (col_idx, pt_type, label, actual_days, interval) in unified:
            col_info[col_idx] = (label, interval, pt_type == 'month', actual_days)

        time_points = {ci: (label, interval, is_month) for ci, (label, interval, is_month, _) in col_info.items()}

        # 读取检测项数据
        data_rows = {}   # item_name → {col_idx: raw_value}
        for r in range(data_start, end_row + 1):
            item_name = ws.cell(r, 2).value
            if item_name and item_name in DETECTION_ITEMS:
                data_rows[item_name] = {c: ws.cell(r, c).value for c in time_points.keys()}

        # 构建记录
        cond_records = []
        for col_idx in sorted(time_points.keys()):
            label, count_val, is_month = time_points[col_idx]
            item_usage = {}        # item_name → numeric usage
            item_checked = {}     # item_name → True if was ✓ (检测不消耗)
            total = 0
            for item_name, col_data in data_rows.items():
                raw = col_data.get(col_idx, None)
                if raw == "✓" or raw == "是":
                    usage = 0
                    item_checked[item_name] = True
                elif raw == "-" or raw is None:
                    usage = 0
                    item_checked[item_name] = False
                else:
                    n = to_num(raw)
                    usage = int(n) if n and n > 0 else 0
                    item_checked[item_name] = False
                item_usage[item_name] = usage
                total += usage

            # 只保留有实际取样量的时间点
            if total > 0:
                # (label, interval, is_month, actual_days) + (item_usage, item_checked, total)
                actual_days = col_info[col_idx][3]
                cond_records.append((label, count_val, item_usage, item_checked, is_month, total, actual_days))

        if cond_records:
            result[cond_name] = cond_records

    return {
        "drug_name": drug_name or plan_path.parent.name,
        "batches": actual_batches,
        "records": result,
    }


# =============================================================================
# 取样计划总表写入
# =============================================================================

def load_master_table():
    if not MASTER_TABLE.exists():
        raise FileNotFoundError(f"取样计划总表不存在: {MASTER_TABLE}")
    return openpyxl.load_workbook(MASTER_TABLE)


def ensure_master_table():
    """
    若取样计划总表不存在，以正确的列结构创建它。
    - 列宽：批号加宽，各日期列可见
    - 冻结：首行 + 首列（项目名称和批号固定）
    - 配色：蓝色表头、交替行斑马纹、状态色彩标注
    """
    global MASTER_TABLE
    if MASTER_TABLE.exists():
        return

    print(f"📄 创建新取样计划总表: {MASTER_TABLE}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "取样计划总表"

    # ---- 列定义 ----
    headers = [
        "项目名称", "批号", "稳定条件", "实际放样日期",
        "时间点", "距放样天数", "计划取样日期", "实际取样日期",
        "状态", "取样量",
        *DETECTION_ITEMS,
    ]

    # 列宽设置（单位：英文字符宽度，1中文≈2单位）
    # 批号最长约 23 字符 → 46；日期列（实际放样日期等 6 字符）→ 18
    col_widths = {
        "项目名称": 20,
        "批号": 30,        # 加宽一倍（默认15 → 30）
        "稳定条件": 14,
        "实际放样日期": 18,
        "时间点": 10,
        "距放样天数": 10,
        "计划取样日期": 18,
        "实际取样日期": 18,
        "状态": 10,
        "取样量": 8,
    }
    for item in DETECTION_ITEMS:
        col_widths[item] = 12

    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci).value = h
        # 设置列宽
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = col_widths.get(h, 12)

    # ---- 冻结首行 + 首列 ----
    ws.freeze_panes = "C2"   # 首行（第1行）+ 前两列（A、B）冻结，B2 为左上角起始

    # 表头样式：深蓝底 + 白字（使用模块级常量）
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(1, ci)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CELL_CENTER
        cell.border = CELL_BORDER

    # 数据行默认样式（写入时动态设置，不预填充空行，避免 max_row 计算错误）

    # ---- 状态列条件格式（待执行蓝/已完成绿/逾期红） ----
    status_col_idx = headers.index("状态") + 1
    col_letter = get_column_letter(status_col_idx)
    status_range = f"{col_letter}2:{col_letter}1000"

    # 待执行：浅蓝底 + 深蓝字
    exec_fill = PatternFill(fill_type="solid", fgColor="DDEEFF")
    exec_font_style = Font(color="1F4E79", size=10)
    exec_dxf = DifferentialStyle(fill=exec_fill, font=Font(color="1F4E79", bold=True))
    ws.conditional_formatting.add(status_range,
        Rule(type="containsText", operator="containsText",
             text="待执行",
             dxf=exec_dxf,
             formula=[f'NOT(ISERROR(SEARCH("待执行",{col_letter}2)))']))

    # 已完成：浅绿底 + 深绿字
    done_fill = PatternFill(fill_type="solid", fgColor="D5F5E3")
    done_dxf = DifferentialStyle(fill=done_fill, font=Font(color="1E8449", bold=True))
    ws.conditional_formatting.add(status_range,
        Rule(type="containsText", operator="containsText",
             text="已完成",
             dxf=done_dxf,
             formula=[f'NOT(ISERROR(SEARCH("已完成",{col_letter}2)))']))

    # 逾期：浅红底 + 深红字
    overdue_fill = PatternFill(fill_type="solid", fgColor="FADBD8")
    overdue_dxf = DifferentialStyle(fill=overdue_fill, font=Font(color="C0392B", bold=True))
    ws.conditional_formatting.add(status_range,
        Rule(type="containsText", operator="containsText",
             text="逾期",
             dxf=overdue_dxf,
             formula=[f'NOT(ISERROR(SEARCH("逾期",{col_letter}2)))']))

    wb.save(MASTER_TABLE)


def header_map(ws):
    """返回 {列名: 列号(1-based)}"""
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def row_exists(ws, h, drug, batch, cond, time_label, actual_days=None):
    for r in range(2, ws.max_row + 1):
        if (ws.cell(r, h["项目名称"]).value == drug
                and ws.cell(r, h["批号"]).value == batch
                and ws.cell(r, h["稳定条件"]).value == cond
                and ws.cell(r, h["时间点"]).value == time_label
                and (actual_days is None or ws.cell(r, h["距放样天数"]).value == actual_days)):
            return r
    return None


def write_to_master(wb, ws, drug_name, batch, cond_name,
                    placement_date, records, notes=""):
    """
    写入记录到取样计划总表。

    列顺序（已按用户要求调整）:
      项目名称 | 批号 | 稳定条件 | 实际放样日期 | 时间点 | 距放样天数 |
      计划取样日期 | 实际取样日期 | 状态 | 取样量 |
      外观/性状 | 有关物质 | pH值 | 不溶性微粒 | 含量 | 无菌 | 细菌内毒素
    """
    h = header_map(ws)
    placement_dt = datetime.strptime(normalize_date(placement_date), "%Y-%m-%d")

    # 确认所有检测项列都存在
    for item in DETECTION_ITEMS:
        if item not in h:
            col_idx = ws.max_column + 1
            ws.cell(1, col_idx).value = item
            h = header_map(ws)   # 重建

    written = 0
    for time_label, count_val, item_usage, item_checked, is_month, total, actual_days in records:
        existing = row_exists(ws, h, drug_name, batch, cond_name, time_label, actual_days)
        if existing:
            print(f"   ⏭️  已存在，跳过: {batch} | {cond_name} | {time_label}")
            continue

        plan_dt = placement_dt + timedelta(days=actual_days)
        row_idx = ws.max_row + 1

        ws.cell(row_idx, h["项目名称"]).value     = drug_name
        ws.cell(row_idx, h["批号"]).value          = batch
        ws.cell(row_idx, h["稳定条件"]).value      = cond_name
        ws.cell(row_idx, h["实际放样日期"]).value   = placement_dt
        ws.cell(row_idx, h["时间点"]).value         = time_label
        ws.cell(row_idx, h["距放样天数"]).value     = actual_days
        ws.cell(row_idx, h["计划取样日期"]).value   = plan_dt
        ws.cell(row_idx, h["实际取样日期"]).value   = None
        ws.cell(row_idx, h["状态"]).value           = "待执行"
        ws.cell(row_idx, h["取样量"]).value         = total

        # 各检测项用量：✓=检测不消耗，数字=消耗量
        for item in DETECTION_ITEMS:
            if item in h:
                if item_checked.get(item, False):
                    ws.cell(row_idx, h[item]).value = "✓"
                else:
                    val = item_usage.get(item, 0) or ""
                    ws.cell(row_idx, h[item]).value = val

        # 设置日期格式
        ws.cell(row_idx, h["实际放样日期"]).number_format  = "YYYY-MM-DD"
        ws.cell(row_idx, h["计划取样日期"]).number_format  = "YYYY-MM-DD"
        ws.cell(row_idx, h["实际取样日期"]).number_format  = "YYYY-MM-DD"

        # 设置斑马纹
        row_fill = ZEBRA_EVEN if row_idx % 2 == 0 else ZEBRA_ODD
        for ci in range(1, len(h) + 1):
            cell = ws.cell(row_idx, ci)
            cell.fill = row_fill
            cell.alignment = CELL_CENTER
            cell.font = DATA_FONT
            cell.border = CELL_BORDER

        # 状态列颜色（直接写死，与条件格式互补）
        status_cell = ws.cell(row_idx, h["状态"])
        sv = str(status_cell.value or "")
        if "待执行" in sv:
            status_cell.fill = STATUS_EXEC[0]
            status_cell.font = STATUS_EXEC[1]
        elif "已完成" in sv:
            status_cell.fill = STATUS_DONE[0]
            status_cell.font = STATUS_DONE[1]
        elif "逾期" in sv:
            status_cell.fill = STATUS_OVER[0]
            status_cell.font = STATUS_OVER[1]

        plan_date = plan_dt.strftime("%Y-%m-%d")
        print(f"   ✅ 写入: {batch} | {cond_name} | {time_label} (+{actual_days}天) → {plan_date} | 消耗{total}支")
        written += 1

    return written


# =============================================================================
# 主入口
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="stability-feedback v5.0.0")
    parser.add_argument("--drug", required=True)
    parser.add_argument("--placement-date", required=True)
    parser.add_argument("--batch",    default="", help="逗号分隔，不指定则处理所有批号")
    parser.add_argument("--condition", default="", help="不指定则处理所有条件")
    parser.add_argument("--notes",    default="")
    parser.add_argument("--plan-file", default="", help="指定计划文件路径（相对于 PLAN_ROOT，或绝对路径），省略则自动查找")
    parser.add_argument("--dry-run",  action="store_true")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"  stability-feedback v5.0.0  放样反馈处理")
    print(f"{'='*60}\n")

    plan_path = resolve_plan_path(args.drug, args.plan_file or None)
    if not plan_path:
        print(f"❌ 未找到药品「{args.drug}」的稳定性研究计划")
        return 1

    batches = [b.strip() for b in args.batch.split(",") if b.strip()] if args.batch else None

    print(f"📄 {plan_path.name}")
    print(f"💊 {args.drug}")
    print(f"📅 实际放样日期: {args.placement_date}")
    if batches:
        print(f"🏷️  批号: {batches}")
    if args.condition:
        print(f"🌡️  条件: {args.condition}")

    parsed = parse_plan(plan_path, batches)
    placement_dt = datetime.strptime(normalize_date(args.placement_date), "%Y-%m-%d")

    print(f"\n{'='*60}")
    print(f"  预览")
    print(f"{'='*60}")

    total_new = 0
    for cond_name, records in parsed["records"].items():
        if args.condition and cond_name != args.condition:
            continue
        if not records:
            continue
        print(f"\n🌡️  {cond_name}")
        print(f"  {'时间点':>8} | {'距放样':>6} | {'计划取样日期':^12} | {'消耗量':>5} | 检测项")
        print(f"  {'-'*8}-+-{'-'*6}-+-{'-'*12}-+-{'-'*5}-+-{'-'*30}")
        # 追踪上一个天数点（用于计算天数点间隔）
        # prev_day_point_actual：上一个天数点的 actual_days
        # prev_month_actual：上一个月的 actual_days（月历点后重置它，供后续天数点作基数）
        prev_day_point_actual = 0   # 上一天数点的 actual_days
        prev_month_actual = 0        # 上一个月的 actual_days
        for time_label, count_val, item_usage, item_checked, is_month, total, actual_days in records:
            plan_dt = placement_dt + timedelta(days=actual_days)
            plan_date = plan_dt.strftime("%Y-%m-%d")
            items_str = ",".join([f"{k}={'✓' if item_checked.get(k, False) else v}"
                                   for k, v in item_usage.items()
                                   if v or item_checked.get(k, False)])
            print(f"  {time_label:>8} | +{actual_days}天 | {plan_date:^12} | {total:>5}支 | {items_str}")
            total_new += 1

    print(f"\n共 {total_new} 条记录 × {len(parsed['batches'])} 批 "
          f"= 约 {total_new * len(parsed['batches'])} 行")
    print(f"（同一批号+条件+时间点已存在则跳过）")

    if args.dry_run:
        print(f"\n🔍 [Dry Run] 未写入")
        return 0

    # 确保总表存在
    ensure_master_table()

    wb = load_master_table()
    ws = wb["取样计划总表"]
    total_written = 0
    for cond_name, records in parsed["records"].items():
        if args.condition and cond_name != args.condition:
            continue
        for batch in parsed["batches"]:
            # records 格式: (label, count_val, item_usage, item_checked, is_month, total)
            records_to_write = list(records)  # 副本，避免 write_to_master 消耗原始列表
            written = write_to_master(wb, ws, parsed["drug_name"], batch,
                                      cond_name, args.placement_date, records_to_write, args.notes)
            total_written += written

    wb.save(MASTER_TABLE)
    print(f"\n✅ 已保存: {MASTER_TABLE}")
    print(f"   新写入: {total_written} 条")
    return 0


if __name__ == "__main__":
    sys.exit(main())
