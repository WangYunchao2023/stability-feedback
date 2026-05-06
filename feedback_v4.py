#!/usr/bin/env python3
"""
stability-feedback v4.0.0 放样反馈处理
========================================
功能：
  1. 接收放样反馈（药品/批号/实际放样日期）
  2. 读取对应药品的稳定性研究计划.xlsx
  3. 解析各条件区块的时间点（含天数和日历月），按规则去重
     - 同一列位置：既有天数又有日历月标注 → 以日历月为准
     - 纯天数列 → 保留
  4. 以实际放样日期为基准，计算各节点计划取样日期
  5. 写入「取样计划总表.xlsx」

用法：
  python feedback_v4.py --drug "黄体酮注射液" \\
    --placement-date "2026-04-09" [--batch "LM-xxx"] [--condition "长期25℃"] \\
    [--notes "无异常"] [--dry-run]
"""

import argparse
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ 需要 openpyxl: pip install openpyxl")
    sys.exit(1)


# ============ 路径配置 ============
AGENT_ROOT = Path("/home/wangyc/.openclaw/workspace/agents/auto-stability-scheme")
PLAN_ROOT = AGENT_ROOT / "stability_plans"
MASTER_TABLE = PLAN_ROOT / "取样计划总表.xlsx"
# =================================

DETECTION_ITEMS = ["外观/性状", "有关物质", "pH值", "不溶性微粒", "含量", "无菌", "细菌内毒素"]


def normalize_date(s: str) -> str:
    if not s:
        return ""
    s = s.strip().replace("年", "-").replace("月", "-").replace("日", "").replace("/", "-")
    parts = re.split(r"[-]", s)
    if len(parts) >= 3:
        return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    return s


def date_to_ms(s: str) -> int:
    if not s:
        return 0
    try:
        return int(datetime.strptime(normalize_date(s), "%Y-%m-%d").timestamp() * 1000)
    except Exception:
        return 0


def find_plan_file(drug_name: str) -> Path:
    for item in PLAN_ROOT.iterdir():
        if item.is_dir() and drug_name in item.name:
            for f in item.glob("*_稳定性计划.xlsx"):
                return f
    return None


def parse_plan(plan_path: Path, target_batches: list = None):
    """
    解析稳定性研究计划.xlsx，返回:
    {
        "drug_name": str,
        "batches": [str],
        "records": {
            "条件名": [(时间点label, 距放样天数, {检测项:用量}, 总计), ...]
        }
    }
    """
    wb = openpyxl.load_workbook(plan_path, data_only=True)
    ws = wb["稳定性研究计划"]

    # ---- 基本信息 ----
    drug_name = ""
    batches = []
    batch_row = None
    for r in range(1, 10):
        for c in range(1, 10):
            v = ws.cell(r, c).value
            if v and "药品" in str(v) and c < 9:
                drug_name = str(ws.cell(r, c + 1).value or "").strip()
            if v and "批号" in str(v):
                batch_row = r
                break
        if batch_row:
            break

    if batch_row:
        # 批号可能在同行多列，也可能在后续行（如第5、6行）
        for r in range(batch_row, min(batch_row + 5, 10)):
            for c in range(1, 10):
                bv = ws.cell(r, c).value
                if bv and "LM" in str(bv):
                    batches.append(str(bv).strip())

    if target_batches:
        batches = [b for b in batches if b in target_batches]

    # ---- 条件区块定义 ----
    # (条件名, header_row, day_row, month_row, data_start, end_row)
    blocks = [
        ("长期25℃", 12, 13, 14, 15, 18),
        ("阴凉20℃", 20, 21, 22, 23, 26),
        ("加速40℃", 28, 29, 30, 31, 34),
        ("高温60℃", 36, 37, 38, 39, 42),
    ]

    result = {}

    for cond_name, hdr_row, day_row_num, month_row_num, data_start, end_row in blocks:
        # 读取 day_row 和 month_row
        day_row = [ws.cell(day_row_num, c).value for c in range(1, 25)]
        month_row = [ws.cell(month_row_num, c).value for c in range(1, 25)]

        # 构建列索引 → 天数 和 列索引 → 月份label
        # 注意：月份和天数是同一列！同列位置，月份标注在month_row，天数在day_row
        col_days = {}    # col_idx(1-based) → int天数
        col_month = {}   # col_idx(1-based) → str "1月"
        col_is_calendar = {}  # col_idx(1-based) → bool 是否日历月

        for col_idx in range(1, 25):
            d_val = day_row[col_idx - 1]
            m_val = month_row[col_idx - 1]

            is_cal = m_val and isinstance(m_val, str) and m_val.endswith("月")
            is_num = isinstance(d_val, (int, float)) and d_val >= 0

            if is_num:
                col_days[col_idx] = int(d_val)
                col_is_calendar[col_idx] = is_cal

            if is_cal:
                # 从月份反推天数
                month_map = {
                    "1月": 30, "2月": 60, "3月": 90, "4月": 120,
                    "6月": 180, "9月": 270, "12月": 360, "15月": 450,
                    "18月": 540, "21月": 630, "24月": 720
                }
                days = month_map.get(m_val, None)
                if days is not None:
                    col_month[col_idx] = m_val
                    if col_idx not in col_days:
                        col_days[col_idx] = days
                        col_is_calendar[col_idx] = True

        # 收集有效时间点（同列位置有天数值的都算，月份覆盖在同位置）
        # 整理：优先取月份label，天数辅助判断
        time_points = {}  # col_idx → (label, days, is_cal)

        for col_idx, days in sorted(col_days.items()):
            if col_idx in col_month:
                time_points[col_idx] = (col_month[col_idx], days, True)
            elif col_idx in col_is_calendar and col_is_calendar[col_idx]:
                # 该列有日历月标注，已在上面处理
                continue
            else:
                # 纯天数列，检查是否需要去重（同一天数是否有日历列）
                is_dup = False
                for c2, days2 in col_days.items():
                    if c2 != col_idx and col_is_calendar.get(c2, False) and days2 == days:
                        is_dup = True
                        break
                if is_dup:
                    continue
                time_points[col_idx] = (f"{days}天", days, False)

        # 读取检测项数据
        data_rows = {}  # item_name → {col_idx: value}
        for r in range(data_start, end_row + 1):
            item_name = ws.cell(r, 2).value
            if item_name and item_name in DETECTION_ITEMS:
                data_rows[item_name] = {c: ws.cell(r, c).value for c in time_points.keys()}

        # 构建记录
        cond_records = []
        for col_idx in sorted(time_points.keys()):
            label, days, is_cal = time_points[col_idx]
            item_usage = {}
            total = 0
            for item_name, col_data in data_rows.items():
                val = col_data.get(col_idx, None)
                if val == "✓" or val == "是":
                    usage = 1
                elif isinstance(val, (int, float)) and val > 0:
                    usage = int(val)
                else:
                    usage = 0
                item_usage[item_name] = usage
                total += usage

            if total > 0:  # 只保留有实际取样量的时间点
                cond_records.append((label, days, item_usage, total))

        if cond_records:
            result[cond_name] = cond_records

    return {
        "drug_name": drug_name or plan_path.parent.name,
        "batches": batches,
        "records": result,
    }


def load_master_table():
    if not MASTER_TABLE.exists():
        raise FileNotFoundError(f"取样计划总表不存在: {MASTER_TABLE}")
    return openpyxl.load_workbook(MASTER_TABLE)


def header_map(ws):
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def row_exists(ws, h, drug, batch, cond, time_label):
    for r in range(2, ws.max_row + 1):
        if (ws.cell(r, h["项目名称"]).value == drug
                and ws.cell(r, h["批号"]).value == batch
                and ws.cell(r, h["稳定条件"]).value == cond
                and ws.cell(r, h["时间点"]).value == time_label):
            return r
    return None


def write_to_master(wb, ws, drug_name, batch, cond_name,
                    placement_date, records, notes=""):
    h = header_map(ws)
    placement_dt = datetime.strptime(normalize_date(placement_date), "%Y-%m-%d")
    placement_ms = date_to_ms(placement_date)

    written = 0
    for time_label, days, item_usage, total in records:
        existing = row_exists(ws, h, drug_name, batch, cond_name, time_label)
        if existing:
            print(f"   ⏭️  已存在，跳过: {batch} | {cond_name} | {time_label}")
            continue

        plan_date = (placement_dt + timedelta(days=days)).strftime("%Y-%m-%d")
        plan_ms = date_to_ms(plan_date)
        row_idx = ws.max_row + 1

        ws.cell(row_idx, h["项目名称"]).value = drug_name
        ws.cell(row_idx, h["批号"]).value = batch
        ws.cell(row_idx, h["稳定条件"]).value = cond_name
        ws.cell(row_idx, h["实际放样日期"]).value = placement_ms
        ws.cell(row_idx, h["时间点"]).value = time_label
        ws.cell(row_idx, h["距放样天数"]).value = days
        ws.cell(row_idx, h["计划取样日期"]).value = plan_ms
        ws.cell(row_idx, h["样品合计"]).value = total
        ws.cell(row_idx, h["状态"]).value = "待执行"
        ws.cell(row_idx, h["备注"]).value = notes or ""

        for item in DETECTION_ITEMS:
            if item in h:
                ws.cell(row_idx, h[item]).value = item_usage.get(item, 0) or ""

        print(f"   ✅ 写入: {batch} | {cond_name} | {time_label} → {plan_date} | {total}支")
        written += 1

    return written


def main():
    parser = argparse.ArgumentParser(description="stability-feedback v4.0.0")
    parser.add_argument("--drug", required=True)
    parser.add_argument("--placement-date", required=True)
    parser.add_argument("--batch", default="", help="逗号分隔，不指定则处理所有批号")
    parser.add_argument("--condition", default="", help="不指定则处理所有条件")
    parser.add_argument("--notes", default="")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"  stability-feedback v4.0.0  放样反馈处理")
    print(f"{'='*60}\n")

    plan_path = find_plan_file(args.drug)
    if not plan_path:
        print(f"❌ 未找到药品「{args.drug}」的稳定性研究计划")
        return 1

    batches = [b.strip() for b in args.batch.split(",") if b.strip()] if args.batch else None

    print(f"📄 {plan_path.name}")
    print(f"💊 {args.drug}")
    print(f"📅 实际放样日期: {args.placement_date}")
    if batches:
        print(f"🏷️  批号: {batches}")
    if args.condition:
        print(f"🌡️  条件: {args.condition}")

    parsed = parse_plan(plan_path, batches)
    placement_dt = datetime.strptime(normalize_date(args.placement_date), "%Y-%m-%d")

    print(f"\n{'='*60}")
    print(f"  预览")
    print(f"{'='*60}")

    total_new = 0
    for cond_name, records in parsed["records"].items():
        if args.condition and cond_name != args.condition:
            continue
        if not records:
            continue
        print(f"\n🌡️  {cond_name}")
        for time_label, days, item_usage, total in records:
            plan_date = (placement_dt + timedelta(days=days)).strftime("%Y-%m-%d")
            items_str = ",".join([k for k, v in item_usage.items() if v])
            print(f"   {time_label:>8} | {plan_date} | +{days}天 | {total}支 | {items_str}")
            total_new += 1

    print(f"\n共 {total_new} 条记录 × {len(parsed['batches'])} 批 = 约{total_new * len(parsed['batches'])}行")
    print(f"（同一批号+条件+时间点已存在则跳过）")

    if args.dry_run:
        print(f"\n🔍 [Dry Run] 未写入")
        return 0

    wb = load_master_table()
    ws = wb["取样计划总表"]
    total_written = 0
    for cond_name, records in parsed["records"].items():
        if args.condition and cond_name != args.condition:
            continue
        for batch in parsed["batches"]:
            written = write_to_master(wb, ws, parsed["drug_name"], batch,
                                      cond_name, args.placement_date, records, args.notes)
            total_written += written

    wb.save(MASTER_TABLE)
    print(f"\n✅ 已保存: {MASTER_TABLE}")
    print(f"   新写入: {total_written} 条")
    return 0


if __name__ == "__main__":
    sys.exit(main())
