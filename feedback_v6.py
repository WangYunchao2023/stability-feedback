#!/usr/bin/env python3
"""
stability-feedback v6.1.0 放样反馈处理
========================================
架构重构（v6.0.0）：
  1. 自动扫描计划文件，发现所有条件区块（不再依赖硬编码 blocks）
  2. Agent 根据每个条件区块的数据特征，判断应该按"月历"还是"天数"计算取样日期
  3. 脚本根据 Agent 判断执行相应逻辑

流程：
  python feedback_v6.py --scan           → 扫描并输出诊断报告（供 Agent 判断用）
  python feedback_v6.py --classify "..."  → 接收 Agent 判断，执行计算和写入
  python feedback_v6.py --drug ...        → 完整流程（scan → Agent 判断 → 执行）
"""

import argparse
import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from dateutil.relativedelta import relativedelta
except ImportError:
    print("❌ 需要 python-dateutil: pip install python-dateutil")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 需要 openpyxl: pip install openpyxl")
    sys.exit(1)


# ============ 路径配置 ============
AGENT_ROOT = Path("/home/wangyc/.openclaw/workspace/agents/auto-formula-scheme")
PLAN_ROOT  = AGENT_ROOT / "stability_plans"
MASTER_TABLE = PLAN_ROOT / "取样计划总表.xlsx"


# ============ 批号规范化 ============
BATCH_PREFIX_ALIAS = {
    "LM002-":   "LMS002-",
    "LM2002-":  "LMS002-",
}


def normalize_batch_key(batch):
    import re
    m = re.match(r'^LM(S)?\d*-(\d{6})-(\d+)$', batch)
    if m:
        date_part = m.group(2)
        seq = m.group(3)
        year2 = date_part[:2]
        month = date_part[2:4]
        day = date_part[4:6]
        return f"{year2}{month}-{day}-{seq}"
    return batch


def normalize_batch_to_actual(batch):
    for plan_prefix, actual_prefix in BATCH_PREFIX_ALIAS.items():
        if batch.startswith(plan_prefix):
            return batch.replace(plan_prefix, actual_prefix, 1)
    return batch


# ============ 检测项目（可扩展）===========
DETECTION_ITEMS = ["外观/性状", "有关物质", "pH值", "不溶性微粒", "含量", "无菌", "细菌内毒素"]

# ============ Excel 样式 ============
_SIDE_THIN = Side(style="thin", color="CCCCCC")
CELL_BORDER = Border(left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=_SIDE_THIN)
CELL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
HDR_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(size=10)
ZEBRA_ODD = PatternFill(fill_type="solid", fgColor="FFFFFF")
ZEBRA_EVEN = PatternFill(fill_type="solid", fgColor="EEF3F8")
STATUS_EXEC = (PatternFill(fill_type="solid", fgColor="DDEEFF"), Font(color="1F4E79", bold=True))
STATUS_DONE = (PatternFill(fill_type="solid", fgColor="D5F5E3"), Font(color="1E8449", bold=True))
STATUS_OVER = (PatternFill(fill_type="solid", fgColor="FADBD8"), Font(color="C0392B", bold=True))


# ============ 工具函数 ============
def normalize_date(s: str) -> str:
    if not s:
        return ""
    s = s.strip().replace("年", "-").replace("月", "-").replace("日", "").replace("/", "-")
    parts = re.split(r"[-]", s)
    if len(parts) >= 3:
        return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    return s


def to_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except (ValueError, AttributeError):
            return None
    return None


def resolve_plan_path(drug_name: str, plan_file_arg: str = None) -> Path:
    if plan_file_arg:
        p = Path(plan_file_arg)
        if p.is_absolute():
            return p
        return PLAN_ROOT / p
    candidates = []
    for item in PLAN_ROOT.iterdir():
        if item.is_dir() and drug_name in item.name:
            for f in item.glob("*_稳定性计划.xlsx"):
                candidates.append(f)
    if not candidates:
        return None
    candidates.sort(key=lambda f: len(f.name), reverse=True)
    return candidates[0]


# =============================================================================
# 核心：自动扫描计划文件，发现所有条件区块
# =============================================================================
def auto_scan_plan(plan_path: Path):
    """
    自动扫描计划文件，发现所有条件区块。
    返回：
    {
        "drug_name": str,
        "batches": [str],
        "blocks": [
            {
                "cond_name": str,
                "header_row": int,
                "time_row": int,       # 天数值所在行
                "month_row": int,      # 月历标签所在行（可能与time_row同一行）
                "data_start": int,
                "data_end": int,
                "day_values": {int: int},   # {列: 天数原始值}
                "month_labels": {int: str}, # {列: "X月"}
                "detection_items": [str],
            },
        ]
    }

    扫描策略：
    - 按行扫描，找到"📦"开头的条件表头
    - 表头下方逐行检查：
        · 有具体数字（天数值）→ 时间点行
        · 有"X月"标签（b2可能为空）→ 月历行
        · 列B为检测项名称 → 检测项数据区
        · 遇到下一个📦或💡 → 区块结束
    """
    wb = openpyxl.load_workbook(plan_path, data_only=True)
    ws = wb["稳定性研究计划"]

    # ---- 药品名 ----
    drug_name = ""
    for r in range(1, 10):
        for c in range(1, 10):
            v = ws.cell(r, c).value
            if v and "药品" in str(v) and c < 9:
                drug_name = str(ws.cell(r, c + 1).value or "").strip()

    # ---- 批号 ----
    batches = []
    batch_row = None
    for r in range(1, 20):
        for c in range(1, 10):
            v = ws.cell(r, c).value
            if v and "批号" in str(v):
                batch_row = r
                break
    if batch_row:
        for r in range(batch_row, batch_row + 25):
            for c in range(1, 10):
                bv = ws.cell(r, c).value
                if bv:
                    bv_str = str(bv).strip()
                    if bv_str.isdigit():
                        continue
                    if "-" not in bv_str and "LM" not in bv_str:
                        continue
                    if re.match(r"^\d{4}-\d{2}-\d{2}$", bv_str):
                        continue
                    # 排除纯 "-" 等无效批号格式
                    if bv_str in ("-", "") or not re.search(r"\d{6}", bv_str):
                        continue
                    batches.append(bv_str)

    actual_batches = [normalize_batch_to_actual(b) for b in batches]

    # ---- 自动发现条件区块 ----
    max_row = ws.max_row
    blocks = []

    for r in range(1, max_row + 1):
        cell_b = ws.cell(r, 2).value
        if not cell_b:
            continue
        cell_str = str(cell_b).strip()
        if not cell_str.startswith("📦"):
            continue

        # 清理条件名称
        cond_name = re.sub(r"[📦💡]|\s*(试验|小计)\s*$", "", cell_str.replace("📦", "")).strip()

        time_row = None
        month_row = None
        data_start = None
        day_values = {}     # {col: day_value int}
        month_labels = {}  # {col: "X月" str}

        # 扫描表头下方最多12行
        for rr in range(r + 1, min(r + 12, max_row + 1)):
            b2 = ws.cell(rr, 2).value
            b2_str = str(b2).strip() if b2 is not None else ""

            # 遇到下一个区块表头或小计行 → 停止
            if b2_str.startswith("📦") or b2_str.startswith("💡") or "计划" in b2_str:
                break

            # 检测项行
            if b2_str in DETECTION_ITEMS:
                if data_start is None:
                    data_start = rr
                continue

            # 扫描该行所有数据列，找数值和月历标签
            row_has_number = False
            row_has_month = False
            row_num_cols = {}
            row_month_cols = {}

            for cc in range(3, 25):
                val = ws.cell(rr, cc).value
                num = to_num(val)
                if num is not None and num >= 0:
                    row_has_number = True
                    row_num_cols[cc] = int(num)
                if isinstance(val, str) and re.search(r"\d+月", val):
                    row_has_month = True
                    row_month_cols[cc] = val

            if row_has_number and data_start is None:
                # 时间点行（第一次出现数值行）
                if time_row is None:
                    time_row = rr
                day_values.update(row_num_cols)
                month_labels.update(row_month_cols)

            elif row_has_month and month_row is None:
                # 月历行（第一次出现月历标签且无天数值）
                month_row = rr
                month_labels.update(row_month_cols)

        # 确定 data_end
        data_end = data_start
        if data_start:
            for rr in range(data_start + 1, r + 20):
                b3 = ws.cell(rr, 2).value
                b3_str = str(b3).strip() if b3 is not None else ""
                if b3_str in DETECTION_ITEMS:
                    data_end = rr
                elif b3_str.startswith("📦") or b3_str.startswith("💡") or "计划" in b3_str:
                    break

        # 检测项列表
        detection_items = []
        if data_start:
            for rr in range(data_start, data_end + 1):
                item = ws.cell(rr, 2).value
                if item and str(item).strip() in DETECTION_ITEMS:
                    detection_items.append(str(item).strip())

        blocks.append({
            "cond_name": cond_name,
            "header_row": r,
            "time_row": time_row,
            "month_row": month_row,
            "data_start": data_start,
            "data_end": data_end,
            "day_values": day_values,
            "month_labels": month_labels,
            "detection_items": detection_items,
        })

    return {
        "drug_name": drug_name or plan_path.parent.name,
        "batches": actual_batches,
        "blocks": blocks,
    }



def format_scan_report(scan_result: dict) -> str:
    """将扫描结果格式化为易读的诊断报告，供 Agent 判断用。"""
    lines = []
    lines.append(f"药品：{scan_result['drug_name']}")
    lines.append(f"批号（共 {len(scan_result['batches'])} 个）：{', '.join(scan_result['batches'][:3])}{'...' if len(scan_result['batches']) > 3 else ''}")
    lines.append(f"")
    lines.append(f"发现 {len(scan_result['blocks'])} 个条件区块：")
    lines.append(f"")
    for i, blk in enumerate(scan_result['blocks']):
        lines.append(f"【区块 {i+1}】{blk['cond_name']}")
        lines.append(f"  表头行：{blk['header_row']} | 时间点行：{blk['time_row']} | 月历行：{blk['month_row']}")
        lines.append(f"  检测项：{', '.join(blk['detection_items']) if blk['detection_items'] else '（未识别）'}")
        # 天数列
        if blk['day_values']:
            day_items = [f"列{c}={v}" for c, v in sorted(blk['day_values'].items())]
            lines.append(f"  天数列（列=值）：{', '.join(day_items)}")
        # 月历列
        if blk['month_labels']:
            month_items = [f"列{c}={v}" for c, v in sorted(blk['month_labels'].items())]
            lines.append(f"  月历列（列=标签）：{', '.join(month_items)}")
        if not blk['day_values'] and not blk['month_labels']:
            lines.append(f"  （无时间点数据）")
        lines.append(f"")
    return "\n".join(lines)


# =============================================================================
# 计算逻辑
# =============================================================================
def compute_time_points(block: dict, calc_type: str, placement_dt: datetime, plan_path: Path):
    """
    根据 calc_type 计算取样时间点。
    calc_type: "month" 或 "day"
    返回 [(label, actual_days, item_usage, item_checked, total), ...]
    """
    # 读取检测项数据
    wb = openpyxl.load_workbook(plan_path, data_only=True)
    ws = wb["稳定性研究计划"]

    time_row = block["time_row"]
    data_start = block["data_start"]
    data_end = block["data_end"]

    # 读取各检测项在各个时间点的原始值
    data_rows = {}  # item_name -> {col: raw_value}
    all_cols = set(list(block["day_values"].keys()) + list(block["month_labels"].keys()))
    for rr in range(data_start, data_end + 1):
        item = ws.cell(rr, 2).value
        if item and str(item).strip() in DETECTION_ITEMS:
            data_rows[str(item).strip()] = {c: ws.cell(rr, c).value for c in all_cols}

    placement_ref = datetime(2026, 4, 10)

    if calc_type == "month":
        # 月历型：
        # 规则：prev_actual 只在纯天数列更新，月历列不更新 prev_actual
        # - 月历列：actual_days = relativedelta(months)，月历列不更新任何状态
        # - 纯天数列（无 month 标签）：actual_days = prev_actual + iv，iv = raw - prev_raw（首个 iv=raw）
        # 这样：月历列之间的 day 列使用上一个月历列的 actual_days 作基数
        month_cols = set(block["month_labels"].keys())
        day_only_cols = set(block["day_values"].keys()) - month_cols

        unified = []
        prev_actual = 0  # 上一个 entry 的 actual_days（供后续天数列作基数）
        prev_raw = 0     # 上一个纯天数列的 raw day 值（供后续天数列 iv 计算用）

        # 月历列：actual_days = relativedelta(months)，不更新 prev_actual 和 prev_raw
        for cc in sorted(block["month_labels"].keys()):
            label = block["month_labels"][cc]
            month_count = int(re.sub(r"\D", "", label))
            actual_days = (placement_ref + relativedelta(months=month_count) - placement_ref).days
            unified.append((cc, label, actual_days))
            # 月历列不更新 prev_raw 和 prev_actual

        # 纯天数列：actual_days = prev_actual + iv（iv = raw - prev_raw，首个 iv=raw）
        for cc in sorted(day_only_cols):
            raw = block["day_values"][cc]
            if raw > 0:
                iv = raw - prev_raw if prev_raw > 0 else raw
                actual_days = prev_actual + iv
                unified.append((cc, str(raw), actual_days))
                prev_raw = raw
                prev_actual = actual_days

        unified.sort(key=lambda x: x[2])

    else:
        # 天数型：只用 day_values，按天数差值累加（忽略 month_labels）
        unified = []
        prev_raw = 0
        prev_actual = 0
        for cc in sorted(block["day_values"].keys()):
            raw = block["day_values"][cc]
            if raw > 0:
                iv = raw - prev_raw if prev_raw > 0 else raw
                actual_days = prev_actual + iv
                unified.append((cc, str(raw), actual_days))
                prev_raw = raw
                prev_actual = actual_days
        unified.sort(key=lambda x: x[2])

    # 构建记录
    records = []
    for cc, label, actual_days in unified:
        item_usage = {}
        item_checked = {}
        total = 0
        for item_name, col_data in data_rows.items():
            raw = col_data.get(cc, None)
            if raw == "✓" or raw == "是":
                usage = 0
                item_checked[item_name] = True
            elif raw == "-" or raw is None:
                usage = 0
                item_checked[item_name] = False
            else:
                n = to_num(raw)
                usage = int(n) if n and n > 0 else 0
                item_checked[item_name] = False
            item_usage[item_name] = usage
            total += usage
        if total > 0:
            records.append((label, actual_days, item_usage, item_checked, total))

    return records


# =============================================================================
# Agent 分类决策（供脚本调用）
# =============================================================================
def agent_classify_blocks(scan_result: dict) -> dict:
    """
    生成 Agent 分类决策报告。
    实际决策由 Agent（外部）根据报告内容自行判断。
    这里只输出格式化的报告内容。
    """
    report = format_scan_report(scan_result)
    decision_prompt = """
请根据上方计划文件的扫描报告，判断每个条件区块应该用哪种取样日期计算方式：

**"month"（月历型）**：取样时间点为"1月""2月"等月历形式，
  按日历月计算（relativedelta），适合长期/阴凉/加速等正式稳定性研究。
  特征：时间点标签是"X月"，月历列和天数列共存。

**"day"（天数型）**：取样时间点为"5天""10天"等纯天数形式，
  直接按天数间隔累加，适合高温/高湿/光照等影响因素试验。
  特征：只有天数字段，无月历标签。

**判断规则**：
- 影响因素试验（高温/高湿/光照/高温加速等）→ day
- 正式稳定性研究（长期/阴凉/加速/中间条件等）→ month
- 不确定时，看数据特征：有月历标签（"X月"）→ month；只有天数 → day

请为每个区块给出判断，格式：
  【区块N】条件名 → month | day

扫描报告：
"""
    return report + "\n" + decision_prompt


# =============================================================================
# 取样计划总表写入
# =============================================================================
def ensure_master_table():
    if MASTER_TABLE.exists():
        return
    print(f"📄 创建新取样计划总表: {MASTER_TABLE}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "取样计划总表"
    headers = [
        "项目名称", "批号", "稳定条件", "实际放样日期",
        "时间点", "距放样天数", "计划取样日期", "实际取样日期",
        "状态", "取样量",
        *DETECTION_ITEMS,
    ]
    col_widths = {
        "项目名称": 20, "批号": 30, "稳定条件": 14,
        "实际放样日期": 18, "时间点": 10, "距放样天数": 10,
        "计划取样日期": 18, "实际取样日期": 18, "状态": 10, "取样量": 8,
    }
    for item in DETECTION_ITEMS:
        col_widths[item] = 12
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci).value = h
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)
    ws.freeze_panes = "C2"
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(1, ci)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CELL_CENTER
        cell.border = CELL_BORDER
    wb.save(MASTER_TABLE)


def load_master_table():
    if not MASTER_TABLE.exists():
        raise FileNotFoundError(f"取样计划总表不存在: {MASTER_TABLE}")
    return openpyxl.load_workbook(MASTER_TABLE)


def header_map(ws):
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def row_exists(ws, h, drug, batch, cond, time_label, actual_days=None):
    for r in range(2, ws.max_row + 1):
        if (ws.cell(r, h["项目名称"]).value == drug
                and ws.cell(r, h["批号"]).value == batch
                and ws.cell(r, h["稳定条件"]).value == cond
                and ws.cell(r, h["时间点"]).value == time_label
                and (actual_days is None or ws.cell(r, h["距放样天数"]).value == actual_days)):
            return r
    return None


def write_to_master(wb, ws, drug_name, batch, cond_name,
                    placement_date, records, notes=""):
    h = header_map(ws)
    placement_dt = datetime.strptime(normalize_date(placement_date), "%Y-%m-%d")
    for item in DETECTION_ITEMS:
        if item not in h:
            col_idx = ws.max_column + 1
            ws.cell(1, col_idx).value = item
            h = header_map(ws)
    written = 0
    for time_label, actual_days, item_usage, item_checked, total in records:
        existing = row_exists(ws, h, drug_name, batch, cond_name, time_label, actual_days)
        if existing:
            print(f"   ⏭️  已存在，跳过: {batch} | {cond_name} | {time_label}")
            continue
        plan_dt = placement_dt + timedelta(days=actual_days)
        row_idx = ws.max_row + 1
        ws.cell(row_idx, h["项目名称"]).value = drug_name
        ws.cell(row_idx, h["批号"]).value = batch
        ws.cell(row_idx, h["稳定条件"]).value = cond_name
        ws.cell(row_idx, h["实际放样日期"]).value = placement_dt
        ws.cell(row_idx, h["时间点"]).value = time_label
        ws.cell(row_idx, h["距放样天数"]).value = actual_days
        ws.cell(row_idx, h["计划取样日期"]).value = plan_dt
        ws.cell(row_idx, h["实际取样日期"]).value = None
        ws.cell(row_idx, h["状态"]).value = "待执行"
        ws.cell(row_idx, h["取样量"]).value = total
        for item in DETECTION_ITEMS:
            if item in h:
                if item_checked.get(item, False):
                    ws.cell(row_idx, h[item]).value = "✓"
                else:
                    val = item_usage.get(item, 0) or ""
                    ws.cell(row_idx, h[item]).value = val
        for col_key in ["实际放样日期", "计划取样日期", "实际取样日期"]:
            ws.cell(row_idx, h[col_key]).number_format = "YYYY-MM-DD"
        row_fill = ZEBRA_EVEN if row_idx % 2 == 0 else ZEBRA_ODD
        for ci in range(1, len(h) + 1):
            cell = ws.cell(row_idx, ci)
            cell.fill = row_fill
            cell.alignment = CELL_CENTER
            cell.font = DATA_FONT
            cell.border = CELL_BORDER
        status_cell = ws.cell(row_idx, h["状态"])
        sv = str(status_cell.value or "")
        if "待执行" in sv:
            status_cell.fill, status_cell.font = STATUS_EXEC
        elif "已完成" in sv:
            status_cell.fill, status_cell.font = STATUS_DONE
        elif "逾期" in sv:
            status_cell.fill, status_cell.font = STATUS_OVER
        print(f"   ✅ 写入: {batch} | {cond_name} | {time_label} (+{actual_days}天) → {plan_dt.strftime('%Y-%m-%d')} | 消耗{total}支")
        written += 1
    return written


# =============================================================================
# 主入口
# =============================================================================
def main():
    parser = argparse.ArgumentParser(description="stability-feedback v6.0.0")
    parser.add_argument("--drug", required=True)
    parser.add_argument("--placement-date", required=True)
    parser.add_argument("--batch", default="", help="逗号分隔，不指定则处理所有批号")
    parser.add_argument("--notes", default="")
    parser.add_argument("--plan-file", default="", help="指定计划文件路径")
    parser.add_argument("--scan", action="store_true", help="仅扫描计划文件，输出诊断报告")
    parser.add_argument("--classify", default="", help="Agent 分类决策 JSON，格式: {\"条件名\": \"month|day\", ...}")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"  stability-feedback v6.0.0  放样反馈处理")
    print(f"{'='*60}\n")

    plan_path = resolve_plan_path(args.drug, args.plan_file or None)
    if not plan_path:
        print(f"❌ 未找到药品「{args.drug}」的计划文件")
        return 1

    print(f"📄 {plan_path.name}")
    print(f"💊 {args.drug}")
    print(f"📅 实际放样日期: {args.placement_date}")

    # ---- 步骤1：扫描 ----
    print(f"\n{'='*60}")
    print(f"  [步骤1] 自动扫描计划文件")
    print(f"{'='*60}")
    scan_result = auto_scan_plan(plan_path)
    print(format_scan_report(scan_result))

    # ---- --scan 模式：只输出报告 ----
    if args.scan:
        print("\n请将上述扫描报告发给 Agent 进行分类判断。")
        print("Agent 返回格式示例：")
        print('  --classify \'{"阴凉20℃": "month", "高温60℃": "day"}\'')
        return 0

    # ---- 步骤2：分类决策 ----
    if not args.classify:
        print("\n❌ 需要 Agent 分类决策。请先运行 --scan，")
        print("   将报告发给 Agent 判断后再执行。")
        example = '{"条件名": "month|day", ...}'
        print(f"   或传入 --classify '{example}'")
        return 1

    try:
        classification = json.loads(args.classify)
    except json.JSONDecodeError:
        print(f"❌ --classify JSON 格式错误: {args.classify}")
        return 1

    print(f"\n{'='*60}")
    print(f"  [步骤2] Agent 分类决策")
    print(f"{'='*60}")
    for cond, calc_type in classification.items():
        print(f"  {cond} → {calc_type}")

    # ---- 步骤3：预览 ----
    placement_dt = datetime.strptime(normalize_date(args.placement_date), "%Y-%m-%d")
    batches = scan_result["batches"]

    print(f"\n{'='*60}")
    print(f"  预览")
    print(f"{'='*60}")

    total_new = 0
    for blk in scan_result["blocks"]:
        cond_name = blk["cond_name"]
        calc_type = classification.get(cond_name, "day")
        records = compute_time_points(blk, calc_type, placement_dt, plan_path)
        if not records:
            continue
        print(f"\n🌡️  {cond_name} [{calc_type}]")
        print(f"  {'时间点':>8} | {'距放样':>6} | {'计划取样日期':^12} | {'消耗量':>5} | 检测项")
        print(f"  {'-'*8}-+-{'-'*6}-+-{'-'*12}-+-{'-'*5}-+-{'-'*30}")
        for label, actual_days, item_usage, item_checked, total in records:
            plan_dt = placement_dt + timedelta(days=actual_days)
            items_str = ",".join([
                f"{k}={'✓' if item_checked.get(k, False) else v}"
                for k, v in item_usage.items()
                if v or item_checked.get(k, False)
            ])
            print(f"  {label:>8} | +{actual_days}天 | {plan_dt.strftime('%Y-%m-%d'):^12} | {total:>5}支 | {items_str}")
            total_new += 1

    print(f"\n共 {total_new} 条记录 × {len(batches)} 批 = 约 {total_new * len(batches)} 行")

    if args.dry_run:
        print(f"\n🔍 [Dry Run] 未写入")
        return 0

    # ---- 步骤4：写入 ----
    ensure_master_table()
    wb = load_master_table()
    ws = wb["取样计划总表"]
    total_written = 0
    for blk in scan_result["blocks"]:
        cond_name = blk["cond_name"]
        calc_type = classification.get(cond_name, "day")
        records = compute_time_points(blk, calc_type, placement_dt, plan_path)
        for batch in batches:
            written = write_to_master(wb, ws, scan_result["drug_name"], batch,
                                      cond_name, args.placement_date, list(records))
            total_written += written

    wb.save(MASTER_TABLE)
    print(f"\n✅ 已保存: {MASTER_TABLE}")
    print(f"   新写入: {total_written} 条")
    return 0


if __name__ == "__main__":
    sys.exit(main())
